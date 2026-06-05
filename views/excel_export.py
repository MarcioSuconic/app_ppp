import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
from models.tarefa import CrudTarefa
from models.equipamento import CrudEquipamento

def gerar_relatorio_excel(caminho_saida=None):
    if not caminho_saida:
        os.makedirs('relatorios', exist_ok=True)
        caminho_saida = f'relatorios/PPP_Relatorio_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # ========== ABA 1: Tarefas ==========
    ws1 = wb.create_sheet("Tarefas")
    cabecalho1 = ["Tarefa", "Função", "Colaborador", "Duração (h)", "Frequência", "Vezes/Mês", "Horas/Mês", "Operação", "Ambiente", "Equipamento", "Observação"]
    ws1.append(cabecalho1)
    
    for t in CrudTarefa.lista_completa_relatorio():
        ws1.append([
            t["tarefa"],
            t["funcao"],
            t["colaborador"],
            round(t["duracao_minutos"] / 60, 1),
            t["frequencia_tipo"],
            round(t["vezes_mes"], 2),
            round(t["horas_mes"], 2),
            t["operacao"] or "-",
            t["ambiente"] or "-",
            t["equipamento"] or "-",
            t["observacao"] or "-"
        ])
    
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    
    # ========== ABA 2: Equipamentos ==========
    ws2 = wb.create_sheet("Equipamentos")
    cabecalho2 = ["Operação", "Ambiente", "Equipamento", "Marca", "Modelo", "Capacidade", "Preço (R$)"]
    ws2.append(cabecalho2)
    
    for eq in CrudEquipamento.listar_todos():
        ws2.append([
            eq.get("operacao_nome") or "-",
            eq.get("ambiente_nome") or "(Uso geral)",
            eq.get("nome") or "-",
            eq.get("marca") or "-",
            eq.get("modelo") or "-",
            eq.get("capacidade") or "-",
            eq.get("preco_estimado", 0)
        ])
    
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    
    # ========== ABA 3: Resumo ==========
    ws3 = wb.create_sheet("Resumo_PP")
    ws3.append(["RESUMO DE HORAS/MÊS POR FUNÇÃO"])
    ws3.append(["Função", "Horas/Mês"])
    
    horas_por_funcao = CrudTarefa.horas_mensais_por_funcao()
    for funcao, horas in horas_por_funcao.items():
        ws3.append([funcao, round(horas, 2)])
    
    ws3.append([])
    ws3.append(["DATAS E INFORMAÇÕES"])
    ws3.append(["Data de Geração", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    
    # ========== ABA 4: Tarefas por Colaborador ==========
    ws4 = wb.create_sheet("Tarefas_por_Colaborador")
    ws4.append(["Colaborador", "Tarefa", "Função", "Duração (h)", "Frequência", "Vezes/Mês", "Horas/Mês", "Operação", "Ambiente", "Equipamento"])
    for item in CrudTarefa.tarefas_por_colaborador():
        ws4.append([item["colaborador"], item["tarefa"], item["funcao"], item["duracao_horas"], item["frequencia"], item["vezes_mes"], item["horas_mes"], item["operacao"], item["ambiente"], item["equipamento"]])
    for cell in ws4[1]: cell.font = Font(bold=True)
    
    # ========== ABA 5: Resumo por Colaborador ==========
    ws5 = wb.create_sheet("Resumo_por_Colaborador")
    ws5.append(["Colaborador", "Total Horas/Mês"])
    for item in CrudTarefa.resumo_horas_por_colaborador():
        ws5.append([item["colaborador"], item["total_horas_mes"]])
    for cell in ws5[1]: cell.font = Font(bold=True)
    
    # ========== ABA 6: Sobre ==========
    ws6 = wb.create_sheet("Sobre_o_PPP")
    for linha in ["PPP - Palmas Project Planner", "Ferramenta de Viabilidade", "", f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"]:
        ws6.append([linha])
    
    for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
        for col in ws.columns:
            max_len = max([len(str(cell.value)) for cell in col if cell.value] or [10])
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    wb.save(caminho_saida)
    return caminho_saida