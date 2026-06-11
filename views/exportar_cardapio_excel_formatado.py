import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os
from views.cardapio_operacao import CardapioOperacao

def gerar_cardapio_excel_formatado(operacao_id, operacao_nome, caminho_saida=None):
    """Gera Excel com cardápio no formato indentado"""
    
    if not caminho_saida:
        os.makedirs('relatorios', exist_ok=True)
        # Remove caracteres inválidos para nome de arquivo
        nome_limpo = operacao_nome.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')
        caminho_saida = f'relatorios/Cardapio_{nome_limpo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Remove caracteres inválidos para título de aba do Excel
    titulo_aba = operacao_nome.replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace(':', '_').replace('[', '_').replace(']', '_')
    ws.title = f"Cardápio - {titulo_aba[:20]}"
    
    # Estilos
    titulo_font = Font(name='Arial', size=16, bold=True)
    categoria_font = Font(name='Arial', size=12, bold=True)
    produto_font = Font(name='Arial', size=11)
    descricao_preco_font = Font(name='Arial', size=10, italic=True)
    
    # Cabeçalho
    ws.merge_cells('A1:B1')
    ws['A1'].value = f'CARDÁPIO - {operacao_nome.upper()}'
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'].value = f'Gerado em: {datetime.now().strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(italic=True, size=9)
    ws.merge_cells('A2:B2')
    
    linha = 4
    
    # ========== PRODUTOS ==========
    produtos = CardapioOperacao.get_produtos_por_operacao(operacao_id)
    if produtos:
        ws.merge_cells(f'A{linha}:B{linha}')
        ws[f'A{linha}'].value = 'PRODUTOS'
        ws[f'A{linha}'].font = Font(size=14, bold=True)
        linha += 1
        
        categorias = {}
        for p in produtos:
            cat = p['categoria']
            if cat not in categorias:
                categorias[cat] = []
            categorias[cat].append(p)
        
        for categoria, itens in categorias.items():
            ws.merge_cells(f'A{linha}:B{linha}')
            ws[f'A{linha}'].value = categoria
            ws[f'A{linha}'].font = categoria_font
            linha += 1
            
            for item in itens:
                ws[f'A{linha}'].value = f'   {item["nome"]}'
                ws[f'A{linha}'].font = produto_font
                linha += 1
                
                descricao = item['descricao'] if item['descricao'] != '-' else ''
                preco = f'R$ {item["preco"]:.2f}'
                
                if descricao:
                    ws[f'A{linha}'].value = f'      {descricao}'
                    ws[f'B{linha}'].value = preco
                else:
                    ws[f'A{linha}'].value = f'      {preco}'
                
                ws[f'A{linha}'].font = descricao_preco_font
                ws[f'B{linha}'].font = descricao_preco_font
                ws[f'B{linha}'].alignment = Alignment(horizontal='right')
                linha += 1
            
            linha += 1
        
        linha += 1
    
    # ========== SERVIÇOS ==========
    servicos = CardapioOperacao.get_servicos_por_operacao(operacao_id)
    if servicos:
        ws.merge_cells(f'A{linha}:B{linha}')
        ws[f'A{linha}'].value = 'SERVIÇOS'
        ws[f'A{linha}'].font = Font(size=14, bold=True)
        linha += 1
        
        categorias = {}
        for s in servicos:
            cat = s['categoria']
            if cat not in categorias:
                categorias[cat] = []
            categorias[cat].append(s)
        
        for categoria, itens in categorias.items():
            ws.merge_cells(f'A{linha}:B{linha}')
            ws[f'A{linha}'].value = categoria
            ws[f'A{linha}'].font = categoria_font
            linha += 1
            
            for item in itens:
                nome = item['nome']
                duracao = f' ({item["duracao"]} min)' if item['duracao'] > 0 else ''
                ws[f'A{linha}'].value = f'   {nome}{duracao}'
                ws[f'A{linha}'].font = produto_font
                linha += 1
                
                descricao = item['descricao'] if item['descricao'] != '-' else ''
                preco = f'R$ {item["preco"]:.2f}'
                
                if descricao:
                    ws[f'A{linha}'].value = f'      {descricao}'
                    ws[f'B{linha}'].value = preco
                else:
                    ws[f'A{linha}'].value = f'      {preco}'
                
                ws[f'A{linha}'].font = descricao_preco_font
                ws[f'B{linha}'].font = descricao_preco_font
                ws[f'B{linha}'].alignment = Alignment(horizontal='right')
                linha += 1
            
            linha += 1
    
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 15
    
    wb.save(caminho_saida)
    return caminho_saida